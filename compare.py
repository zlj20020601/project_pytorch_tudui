import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# 加载子表
df_si1 = pd.read_excel('E:\\shijian\\1.xlsx', sheet_name='si1')
df_si = pd.read_excel('E:\\shijian\\1.xlsx', sheet_name='si')

# 确定共同列
common_columns = df_si1.columns.intersection(df_si.columns)

# 仅比较共同列
df_si1_common = df_si1[common_columns]
df_si_common = df_si[common_columns]

# 对比共同列并标记差异
def safe_convert(val):
    try:
        # 如果是数值型数据，则返回其浮动数值
        return float(val) if isinstance(val, (int, float, str)) and val != '' else val
    except ValueError:
        # 如果无法转换为数值型，则返回原始值
        return val

# 将每一列的值进行安全转换
df_si1_common_float = df_si1_common.applymap(safe_convert)
df_si_common_float = df_si_common.applymap(safe_convert)

# 比较浮动数值，忽略非数值类型
df_comparison = df_si1_common_float != df_si_common_float
df_comparison = df_comparison.applymap(lambda x: '不同' if x else '相同')

# 构建完整的比较结果DataFrame
df_comparison_full = pd.concat([df_si1_common, df_si_common, df_comparison], axis=1)
df_comparison_full.columns = [f'si1_{col}' for col in common_columns] + [f'si_{col}' for col in common_columns] + [f'diff_{col}' for col in common_columns]

# 保存比较结果到一个新的 Excel 文件
output_file_path = 'E:\\shijian\\compare.xlsx'
df_comparison_full.to_excel(output_file_path, index=False)

# 使用 openpyxl 打开结果文件并进行标红
wb = openpyxl.load_workbook(output_file_path)
ws = wb.active

# 设置红色填充样式
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# 标记差异的单元格为红色
for row in range(2, ws.max_row + 1):  # 从第二行开始，因为第一行是表头
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        # 如果该单元格显示为'不同'
        if cell.value == '不同':
            cell.fill = red_fill

# 保存修改后的 Excel 文件
wb.save(output_file_path)

print(f'比较结果已保存并标红至：{output_file_path}')
